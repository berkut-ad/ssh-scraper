import os
import re
import difflib
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

LOG_DIR = 'logs'
OUTPUT_FILE = 'command_diff_report.xlsx'

# Utility: parse log file and extract {command: output}
def parse_log_file(filepath):
    with open(filepath, 'r') as f:
        content = f.read()
    matches = re.split(r"\n>> (.+?)\n", content)[1:]  # Skip header text
    commands = {}
    for i in range(0, len(matches), 2):
        cmd = matches[i].strip()
        output = matches[i + 1].strip()
        commands[cmd] = output
    return commands

# Utility: do line-by-line diff highlighting
def diff_lines(ref_output, target_output):
    ref_lines = ref_output.splitlines()
    target_lines = target_output.splitlines()
    diff = difflib.ndiff(ref_lines, target_lines)
    highlighted = []
    for line in diff:
        if line.startswith("+ "):
            highlighted.append(("ADDED", line[2:]))
        elif line.startswith("- "):
            highlighted.append(("REMOVED", line[2:]))
        elif line.startswith("? "):
            continue  # Skip hints
        else:
            highlighted.append(("SAME", line[2:]))
    return highlighted

# Load all log files
device_files = [f for f in os.listdir(LOG_DIR) if f.endswith(".txt")]
device_outputs = {}
for fname in device_files:
    ip = fname.split("_")[0]
    filepath = os.path.join(LOG_DIR, fname)
    device_outputs[ip] = parse_log_file(filepath)

# Invert structure: {command: {ip: output}}
command_data = defaultdict(dict)
for ip, commands in device_outputs.items():
    for cmd, output in commands.items():
        command_data[cmd][ip] = output

# Excel workbook creation
wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

header_font = Font(bold=True)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
fill_diff = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")

for cmd, device_map in command_data.items():
    sheet_title = cmd[:31] if len(cmd) > 31 else cmd  # Excel sheet name max = 31 chars
    ws = wb.create_sheet(title=sheet_title)

    # Build header row: first cell is "Command", rest are IPs
    ips_sorted = sorted(device_map.keys())
    header_row = ["Command"] + ips_sorted
    ws.append(header_row)
    for col_idx, val in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col_idx, value=val)
        cell.font = header_font
        cell.alignment = wrap_alignment

    # Prepare the output row
    output_row = [cmd]  # First column is the command itself
    ref_ip = ips_sorted[0]
    ref_output = device_map[ref_ip]

    for ip in ips_sorted:
        output = device_map.get(ip, "")
        if ip == ref_ip:
            output_row.append(output)
        else:
            diffs = diff_lines(ref_output, output)
            colored_lines = [line for status, line in diffs]
            final_output = "\n".join(colored_lines)
            output_row.append(final_output)

    # Add the row
    ws.append(output_row)

    # Style and highlight if needed
    for col_idx, ip in enumerate(ips_sorted, start=2):
        cell = ws.cell(row=2, column=col_idx)
        cell.alignment = wrap_alignment
        if device_map.get(ip) and ip != ref_ip:
            diffs = diff_lines(ref_output, device_map[ip])
            if any(status != "SAME" for status, _ in diffs):
                cell.fill = fill_diff

    # Auto column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 5, 100)

# Save final Excel
wb.save(OUTPUT_FILE)
print(f"Saved: {OUTPUT_FILE}")
